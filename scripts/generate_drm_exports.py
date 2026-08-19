import json
import re
import zipfile
from copy import deepcopy
from html import escape
from pathlib import Path
from xml.sax.saxutils import escape as xesc

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter


ROOT = Path(__file__).resolve().parents[1]
OUT = ROOT / "exports"
CURRENT_XLSX = OUT / "Current_Previous_Year_PU_Demand_Analysis.xlsx"
CURRENT_PDF = OUT / "Current_Previous_Year_PU_Demand_Analysis.pdf"
FR_XLSX = OUT / "FR_Budget_Status.xlsx"
FR_PDF = OUT / "FR_Budget_Status.pdf"
CURRENT_PPTX = OUT / "Moradabad_Division_Current_Year_Budget_Analysis.pptx"
PPTX = OUT / "Moradabad_Division_DRM_Budget_FR_Analysis.pptx"
XLSX = OUT / "Moradabad_Division_DRM_Budget_FR_Analysis.xlsx"
TEMPLATE_CANDIDATES = [
    Path(r"C:\Users\HP\Dropbox\Revenue PU Laibilities\PPT PORTAL\Moradabad Division Quarty FR and Revenue Budget Analysis DRM.pptx"),
    Path(r"C:\Users\HP\Dropbox\Revenue PU Laibilities\PPT PORTAL\Moradabad Division Quarty FR and Revenue Budget Analysis DRM JULY.pptx"),
    Path(r"C:\Users\HP\Dropbox\Revenue PU Laibilities\PPT PORTAL\Moradabad Division Quarty FR and Revenue Budget Analysis DRM JUN.pptx"),
    Path(r"C:\Users\HP\Dropbox\Revenue PU Laibilities\PPT PORTAL\Moradabad Division Quarty FR and Revenue Budget Analysis.pptx"),
]
TEMPLATE_PPTX = next((path for path in TEMPLATE_CANDIDATES if path.exists()), TEMPLATE_CANDIDATES[0])
SLIDE_W, SLIDE_H = 12192000, 6858000
BLUE = "1F4E79"
NAVY = "003366"
LIGHT = "E8F2F8"
YELLOW = "FFF2CC"
WHITE = "FFFFFF"
BLACK = "000000"
GREEN = "25A55B"
AMBER = "F2C230"
RED = "D92323"
GREEN_LIGHT = "DFF3E7"
AMBER_LIGHT = "FFF2CC"
RED_LIGHT = "FCE4E4"


def load_json_assignment(path, name):
    text = path.read_text(encoding="utf-8")
    match = re.search(rf"{re.escape(name)}\s*=\s*(.*?);?\s*$", text, re.S)
    if not match:
        raise RuntimeError(f"Cannot locate {name} in {path}")
    return json.loads(match.group(1))


def load_fr_data():
    text = (ROOT / "pages" / "fr.html").read_text(encoding="utf-8")
    match = re.search(r"const\s+workbookData\s*=\s*(\[.*?\]);\s*const\s+funds", text, re.S)
    if not match:
        raise RuntimeError("Cannot locate FR workbookData")
    return json.loads(match.group(1))


def load_fr_as_on():
    text = (ROOT / "pages" / "fr.html").read_text(encoding="utf-8")
    match = re.search(r"Data as on <strong>(.*?)</strong>", text)
    return clean_text(match.group(1)) if match else "FR as uploaded"


def current_as_on_label():
    text = (ROOT / "data" / "current_payload.js").read_text(encoding="utf-8")
    match = re.search(r"window\.CURRENT_PAYLOAD_META\s*=\s*(\{.*?\});", text, re.S)
    meta = json.loads(match.group(1)) if match else {}
    return f"{meta.get('completedMonth', 'Completed month')} | uploaded {meta.get('statusAsOn', '')}"


def inr(value, decimals=0):
    try:
        n = float(value or 0)
    except Exception:
        return ""
    return f"{n:,.{decimals}f}" if decimals else f"{round(n):,}"


def money(value):
    try:
        n = float(value or 0)
    except Exception:
        n = 0
    return f"{inr(n)}\nCr. {n / 10000:.2f}"


def pct(value):
    try:
        return f"{round(float(value or 0))}"
    except Exception:
        return "0"


def row_name(row):
    return row.get("Name") or row.get("PU") or row.get("Demand") or ""


def is_total(row):
    return row_name(row).strip().lower() == "total"


def number_value(value):
    try:
        return float(value or 0)
    except Exception:
        return 0


def code_from_label(label, prefix):
    match = re.search(rf"{re.escape(prefix)}\s*-\s*([0-9A-Z]+)", str(label or ""), re.I)
    return match.group(1).upper() if match else ""


def demand_key(label):
    match = re.search(r"Demand\s+([0-9A-Z]+)", str(label or ""), re.I)
    return match.group(1).upper() if match else ""


def is_demand_suspense(row):
    name = row_name(row).upper()
    department = str(row.get("Department") or "").upper()
    return bool(re.search(r"\b(12N|10N)\b", name)) or "SUSPENSE" in department


def detail_rows(rows):
    return [row for row in rows or [] if not is_total(row)]


def normal_total_rows(rows):
    return [row for row in detail_rows(rows) if not is_demand_suspense(row)]


def demand_suspense_rows(rows):
    return [row for row in detail_rows(rows) if is_demand_suspense(row)]


def latest_report_year(reports, offset=0):
    years = reports.get("years") or []
    idx = max(0, len(years) - 1 - offset)
    return (years[idx] or {}).get("fy", "")


def match_monthly_key(scope, label, bucket):
    keys = list((bucket or {}).keys())
    if label in bucket:
        return label
    if scope == "pu":
        code = code_from_label(label, "PU")
        return next((key for key in keys if code_from_label(key, "PU") == code), "")
    if scope == "demand":
        demand = demand_key(label)
        smh_match = re.search(r"/\s*([0-9A-Z]+)", str(label or ""), re.I)
        smh = smh_match.group(1).upper() if smh_match else ""
        return next((key for key in keys if demand_key(key) == demand and (not smh or f"SMH {smh}" in key.upper())), "")
    return next((key for key in keys if key == label), "")


def month_actual(reports, scope, label, fy, count):
    bucket = ((reports.get("monthly") or {}).get(scope) or {})
    key = match_monthly_key(scope, label, bucket)
    values = (bucket.get(key) or {}).get(fy) if key else None
    if not isinstance(values, list):
        return None
    return sum(number_value(value) for value in values[:count])


def relabel_period(text, month="JUN", year=2026, count=3):
    label = f"{month} {year}"
    return (str(text or "")
        .replace("JUL 2026", label)
        .replace("JUL 2025", f"{month} 2025")
        .replace("/ 12 * 4", f"/ 12 * {count}")
        .replace("BP UPTO JUL 2026", f"BP UPTO {label}"))


def summary_row(label, oba, ae, months=3, bp_override=None):
    bp = number_value(bp_override) if bp_override is not None else oba / 12 * months
    return {
        "Name": label,
        "OBA": oba,
        "BP": bp,
        "AE": ae,
        "Variation": ae - bp,
        "BPPercent": ae / bp * 100 if bp else 0,
        "Remaining": oba - ae,
        "OBAPercent": ae / oba * 100 if oba else 0,
        "Months": months,
    }


def add_total(rows):
    normal = normal_total_rows(rows)
    suspense = demand_suspense_rows(rows)
    months = number_value((normal[0] if normal else rows[0] if rows else {}).get("Months") or 3)
    oba = sum(number_value(row.get("OBA")) for row in normal)
    ae = sum(number_value(row.get("AE")) for row in normal)
    bp = sum(number_value(row.get("BP")) for row in normal)
    return normal + [summary_row("Total", oba, ae, months, bp)] + suspense


def filtered_pu_rows(rows, codes):
    wanted = {str(code).zfill(2) for code in codes}
    detail = [row for row in detail_rows(rows) if code_from_label(row_name(row), "PU").zfill(2) in wanted]
    return add_total(detail)


def apply_completed_period(payload):
    reports = json.loads((ROOT / "data" / "reports-data.json").read_text(encoding="utf-8-sig"))
    fy = latest_report_year(reports)
    view = deepcopy(payload)
    period = {"month": "JUL", "year": 2026, "count": 4, "label": "JUL 2026"}
    for key in ("demand", "staff", "nonstaff"):
        tab = view.get(key)
        if not tab or not tab.get("rows"):
            continue
        scope = "demand" if key == "demand" else "pu"
        rows = []
        for row in detail_rows(tab["rows"]):
            next_row = dict(row)
            actual = month_actual(reports, scope, row_name(row), fy, period["count"])
            if actual is not None:
                next_row["AE"] = actual
            next_row["Months"] = period["count"]
            # Keep the portal/source BP proportion instead of recomputing a flat monthly ratio.
            # Source BP can include budget-distribution logic that is not always OBA / 12 * months.
            next_row["BP"] = number_value(next_row.get("BP"))
            next_row["Variation"] = number_value(next_row.get("AE")) - number_value(next_row.get("BP"))
            next_row["BPPercent"] = number_value(next_row.get("AE")) / number_value(next_row.get("BP")) * 100 if number_value(next_row.get("BP")) else 0
            next_row["Remaining"] = number_value(next_row.get("OBA")) - number_value(next_row.get("AE"))
            next_row["OBAPercent"] = number_value(next_row.get("AE")) / number_value(next_row.get("OBA")) * 100 if number_value(next_row.get("OBA")) else 0
            rows.append(next_row)
        tab["columns"] = [{**col, "label": relabel_period(col.get("label"), period["month"], period["year"], period["count"])} for col in tab.get("columns", [])]
        tab["title"] = f'{tab.get("title", "")} - Completed Month Projection - July 2026 (04 months)'
        tab["rows"] = add_total(rows)
    return view


def utilization_class(value):
    try:
        v = float(value or 0)
    except Exception:
        v = 0
    if v < 0 or v > 100:
        return "red"
    if v >= 75:
        return "amber"
    return "green"


def utilization_fill(value, light=False):
    cls = utilization_class(value)
    if light:
        return {"green": GREEN_LIGHT, "amber": AMBER_LIGHT, "red": RED_LIGHT}.get(cls, GREEN_LIGHT)
    return {"green": GREEN, "amber": AMBER, "red": RED}.get(cls, GREEN)


def utilization_text_color(value):
    return BLACK if utilization_class(value) == "amber" else WHITE


def utilization_dot(value):
    return {"green": "G", "amber": "A", "red": "R"}.get(utilization_class(value), "G")


def display_cell(row, col):
    key, fmt = col["key"], col.get("format")
    val = row.get(key, "")
    if fmt == "money":
        return money(val)
    if key in ("BPPercent", "OBAPercent"):
        return pct(val)
    if fmt == "int":
        return pct(val)
    return str(val)


def table_from_payload(tab, columns=None, rows=None):
    columns = columns or tab["columns"]
    rows = rows or tab["rows"]
    headers = [c["label"] for c in columns]
    body = [[display_cell(r, c) for c in columns] for r in rows]
    return headers, body


def fr_report_table(sheet):
    headers = ["Plan Head", "Plan Head Name", "SBA 2026-27", "AE", "Variation (AE - SBA)", "% SBA"]
    body = []
    for rec in sheet["records"] + [sheet["total"]]:
        total = rec["funds"]["TOTAL"]
        body.append([
            rec.get("planHead") or "Total",
            rec.get("planName") or "",
            money(total["sba"]),
            money(total["ae"]),
            money(total["ae"] - total["sba"]),
            f"{total.get('spentPct', 0):.2f}",
        ])
    return headers, body


def fr_fund_table(sheet):
    headers = ["Fund", "SBA", "AE", "Available", "Expensed %", "Remaining %"]
    body = []
    for fund, values in sheet["fundAnalysis"].items():
        body.append([fund, money(values["sba"]), money(values["ae"]), money(values["available"]), f"{values.get('spentPct',0):.2f}", f"{values.get('remainPct',0):.2f}"])
    return headers, body


def style_ws(ws):
    black = Side(style="thin", color="000000")
    border = Border(left=black, right=black, top=black, bottom=black)
    for row in ws.iter_rows():
        for cell in row:
            cell.border = border
            cell.alignment = Alignment(wrap_text=True, vertical="center")
            cell.font = Font(name="Times New Roman", size=10)
            if cell.row == 1:
                cell.font = Font(name="Times New Roman", size=14, bold=True, color=BLUE)
            elif cell.row == 2:
                cell.font = Font(name="Times New Roman", size=10, bold=True, color="607080")
            elif cell.row == 4:
                cell.fill = PatternFill("solid", fgColor=BLUE)
                cell.font = Font(name="Times New Roman", size=10, bold=True, color=WHITE)
            elif cell.row % 2 == 0 and cell.row > 4:
                cell.fill = PatternFill("solid", fgColor=LIGHT)
    for col in ws.columns:
        letter = get_column_letter(col[0].column)
        width = max(len(str(c.value or "").split("\n")[0]) for c in col) + 3
        ws.column_dimensions[letter].width = min(max(width, 10), 28)
    ws.page_margins.left = ws.page_margins.right = ws.page_margins.top = ws.page_margins.bottom = 0.25
    ws.page_setup.orientation = "landscape"
    ws.sheet_properties.pageSetUpPr.fitToPage = True
    ws.page_setup.fitToWidth = 1


def write_excel(sections, output_path=XLSX):
    wb = Workbook()
    wb.remove(wb.active)
    for title, headers, body in sections:
        ws = wb.create_sheet(re.sub(r"[\\/?*\[\]:]", " ", title)[:31])
        ws.append([title])
        ws.append(["Figures in '000 with Crore shown below. Generated from latest portal data."])
        ws.append([])
        ws.append(headers)
        for row in body:
            ws.append(row)
        style_ws(ws)
    output_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(output_path)


def pdf_escape(value):
    return clean_text(value).replace("\\", "\\\\").replace("(", "\\(").replace(")", "\\)")


def pdf_text_at(x, y, text, size=10, bold=False):
    font = "F2" if bold else "F1"
    return f"BT /{font} {size} Tf {x:.2f} {y:.2f} Td ({pdf_escape(text)}) Tj ET\n"


def pdf_rect(x, y, w, h, fill=None, stroke=True):
    commands = []
    if fill:
        r = int(fill[0:2], 16) / 255
        g = int(fill[2:4], 16) / 255
        b = int(fill[4:6], 16) / 255
        commands.append(f"{r:.3f} {g:.3f} {b:.3f} rg\n")
    commands.append(f"{x:.2f} {y:.2f} {w:.2f} {h:.2f} re ")
    if fill and stroke:
        commands.append("B\n")
    elif fill:
        commands.append("f\n")
    else:
        commands.append("S\n")
    return "".join(commands)


def pdf_color(color):
    r = int(color[0:2], 16) / 255
    g = int(color[2:4], 16) / 255
    b = int(color[4:6], 16) / 255
    return f"{r:.3f} {g:.3f} {b:.3f} rg\n"


def shorten_for_pdf(value, limit):
    text = clean_text(value).replace("\n", " / ")
    return text if len(text) <= limit else text[: max(0, limit - 1)] + "."


def pdf_column_widths(headers, page_width):
    weights = column_weights(headers)
    return [page_width * weight for weight in weights]


def pdf_table_pages(title, headers, rows):
    page_w, page_h, margin = 842, 595, 18
    table_w = page_w - margin * 2
    col_widths = pdf_column_widths(headers, table_w)
    row_h = 24
    header_h = 30
    max_rows = int((page_h - 95 - header_h) // row_h)
    pages = []
    for start in range(0, max(len(rows), 1), max_rows):
        part = rows[start:start + max_rows]
        y = page_h - margin - 18
        content = "0 0 0 RG 0.4 w\n"
        content += pdf_color(BLUE)
        content += pdf_text_at(margin, y, title if start == 0 else f"{title} continued", 15, True)
        y -= 18
        content += pdf_color("666666")
        content += pdf_text_at(margin, y, "Figures in '000 with Crore equivalent below where applicable. Generated from latest portal data.", 8)
        y -= 22
        x = margin
        for idx, header in enumerate(headers):
            w = col_widths[idx]
            content += pdf_rect(x, y - header_h + 4, w, header_h, BLUE)
            content += pdf_color(WHITE)
            content += pdf_text_at(x + 2, y - 9, shorten_for_pdf(header, max(6, int(w / 4.2))), 7, True)
            x += w
        y -= header_h
        for ridx, row in enumerate(part):
            first = clean_text(row[0]).strip().lower() if row else ""
            fill = "D9EAF7" if ridx % 2 else WHITE
            if first == "total":
                fill = "C8D6E8"
            x = margin
            for cidx, value in enumerate(row):
                w = col_widths[cidx]
                cell_fill = fill
                text_color = BLACK
                if is_percent_header(headers[cidx]) and clean_text(value).strip():
                    cell_fill = utilization_fill(numeric_prefix(value), light=(first != "total"))
                    text_color = BLACK if first != "total" else utilization_text_color(numeric_prefix(value))
                content += pdf_rect(x, y - row_h + 4, w, row_h, cell_fill)
                content += pdf_color(text_color)
                limit = max(5, int(w / 4.7))
                content += pdf_text_at(x + 2, y - 8, shorten_for_pdf(value, limit), 7.2, first == "total")
                x += w
            y -= row_h
        pages.append((page_w, page_h, content))
    return pages


def write_pdf(sections, output_path):
    pages = []
    for title, headers, body in sections:
        pages.extend(pdf_table_pages(title, headers, body))
    objects = [
        "<< /Type /Catalog /Pages 2 0 R >>",
        "",
    ]
    page_refs = []
    next_obj = 3
    for page_w, page_h, content in pages:
        content_bytes = content.encode("latin-1", errors="replace")
        page_obj = next_obj
        content_obj = next_obj + 1
        page_refs.append(f"{page_obj} 0 R")
        objects.append(f"<< /Type /Page /Parent 2 0 R /MediaBox [0 0 {page_w} {page_h}] /Resources << /Font << /F1 {len(pages) * 2 + 3} 0 R /F2 {len(pages) * 2 + 4} 0 R >> >> /Contents {content_obj} 0 R >>")
        objects.append(f"<< /Length {len(content_bytes)} >>\nstream\n{content}\nendstream")
        next_obj += 2
    objects[1] = f"<< /Type /Pages /Count {len(page_refs)} /Kids [{' '.join(page_refs)}] >>"
    objects.append("<< /Type /Font /Subtype /Type1 /BaseFont /Times-Roman >>")
    objects.append("<< /Type /Font /Subtype /Type1 /BaseFont /Times-Bold >>")
    output_path.parent.mkdir(parents=True, exist_ok=True)
    offsets = []
    data = b"%PDF-1.4\n"
    for idx, obj in enumerate(objects, 1):
        offsets.append(len(data))
        data += f"{idx} 0 obj\n".encode("latin-1")
        data += obj.encode("latin-1", errors="replace")
        data += b"\nendobj\n"
    xref = len(data)
    data += f"xref\n0 {len(objects) + 1}\n0000000000 65535 f \n".encode("latin-1")
    for offset in offsets:
        data += f"{offset:010d} 00000 n \n".encode("latin-1")
    data += f"trailer << /Size {len(objects) + 1} /Root 1 0 R >>\nstartxref\n{xref}\n%%EOF\n".encode("latin-1")
    output_path.write_bytes(data)


def clean_text(value):
    text = str(value if value is not None else "")
    replacements = {"\u2013": "-", "\u2014": "-", "\u2011": "-", "\u00a0": " "}
    for src, dst in replacements.items():
        text = text.replace(src, dst)
    return "".join(ch if ch == "\n" or 32 <= ord(ch) <= 126 else " " for ch in text)


def text_runs(text, size=900, bold=False, color=BLACK):
    bold_attr = ' b="1"' if bold else ""
    runs = []
    for idx, part in enumerate(clean_text(text).split("\n")):
        if idx:
            runs.append("<a:br/>")
        runs.append(f'<a:r><a:rPr lang="en-US" sz="{size}"{bold_attr}><a:solidFill><a:srgbClr val="{color}"/></a:solidFill><a:latin typeface="Times New Roman"/></a:rPr><a:t>{xesc(part)}</a:t></a:r>')
    return "".join(runs)


def ppt_text_box(shape_id, x, y, w, h, text, size=1200, bold=False, fill=None, color=BLACK, align="ctr"):
    fill_xml = f'<a:solidFill><a:srgbClr val="{fill}"/></a:solidFill>' if fill else "<a:noFill/>"
    return f'''<p:sp><p:nvSpPr><p:cNvPr id="{shape_id}" name="Text {shape_id}"/><p:cNvSpPr/><p:nvPr/></p:nvSpPr><p:spPr><a:xfrm><a:off x="{x}" y="{y}"/><a:ext cx="{w}" cy="{h}"/></a:xfrm><a:prstGeom prst="rect"><a:avLst/></a:prstGeom>{fill_xml}<a:ln><a:noFill/></a:ln></p:spPr><p:txBody><a:bodyPr wrap="square" anchor="ctr" lIns="50000" rIns="50000" tIns="25000" bIns="25000"/><a:lstStyle/><a:p><a:pPr algn="{align}"/>{text_runs(text, size, bold, color)}</a:p></p:txBody></p:sp>'''


def tc_pr(fill):
    border = f'<a:solidFill><a:srgbClr val="{BLACK}"/></a:solidFill>'
    lines = "".join(f'<a:ln{side} w="12700">{border}</a:ln{side}>' for side in ["L", "R", "T", "B"])
    return f'<a:tcPr marL="12000" marR="12000" marT="8000" marB="8000"><a:solidFill><a:srgbClr val="{fill}"/></a:solidFill>{lines}</a:tcPr>'


def table_cell(text, fill, size, bold=False, color=BLACK, align="ctr"):
    anchor = "ctr"
    return f'''<a:tc><a:txBody><a:bodyPr wrap="square" anchor="{anchor}" lIns="18000" rIns="18000" tIns="10000" bIns="10000"/><a:lstStyle/><a:p><a:pPr algn="{align}"/>{text_runs(text, size, bold, color)}</a:p></a:txBody>{tc_pr(fill)}</a:tc>'''


def column_weights(headers):
    weights = []
    has_department = len(headers) > 1 and "department" in clean_text(headers[1]).lower()
    for index, header in enumerate(headers):
        label = clean_text(header).lower()
        if index == 0:
            weights.append(1.35 if has_department else 1.65)
        elif "department" in label or "name" in label:
            weights.append(2.05)
        elif "%" in label or "percent" in label:
            weights.append(0.72)
        elif "variation" in label or "remaining" in label:
            weights.append(1.18)
        elif "oba" in label or "bp" in label or "actual" in label or "ae" in label or "sba" in label:
            weights.append(1.08)
        else:
            weights.append(1.05)
    total = sum(weights) or 1
    return [weight / total for weight in weights]


def numeric_prefix(value):
    match = re.search(r"-?\d+(?:\.\d+)?", clean_text(value).replace(",", ""))
    return float(match.group(0)) if match else 0


def is_percent_header(header):
    label = clean_text(header).lower()
    return "%" in label or "expensed" in label or "utilized" in label


def ppt_table(shape_id, x, y, w, h, headers, rows):
    col_count = len(headers)
    row_count = len(rows) + 1
    row_h = int(h / max(row_count, 1))
    header_size = 1080 if col_count <= 8 else 1000
    body_size = 1000
    col_widths = [int(w * weight) for weight in column_weights(headers)]
    col_widths[-1] += int(w) - sum(col_widths)
    grid = "".join(f'<a:gridCol w="{col_w}"/>' for col_w in col_widths)
    trs = [
        f'<a:tr h="{row_h}">' + "".join(table_cell(head, BLUE, header_size, True, WHITE, "ctr") for head in headers) + "</a:tr>"
    ]
    percent_cols = {idx for idx, head in enumerate(headers) if is_percent_header(head)}
    for r_idx, row in enumerate(rows):
        first = clean_text(row[0]).strip().lower() if row else ""
        base_fill = "C8D6E8" if first == "total" else ("D9EAF7" if r_idx % 2 else WHITE)
        cells = []
        for c_idx, value in enumerate(row):
            align = "l" if c_idx == 1 and ("department" in clean_text(headers[c_idx]).lower() or "name" in clean_text(headers[c_idx]).lower()) else "ctr"
            bold = first == "total" or c_idx in (0, col_count - 1) or c_idx in percent_cols
            size = body_size + 130 if first == "total" and col_count <= 9 else body_size
            fill = base_fill
            color = BLACK
            if c_idx in percent_cols and clean_text(value).strip():
                n = numeric_prefix(value)
                fill = utilization_fill(n, light=(first != "total"))
                color = BLACK if first != "total" else utilization_text_color(n)
            cells.append(table_cell(value, fill, size, bold, color, align))
        trs.append(f'<a:tr h="{row_h}">' + "".join(cells) + "</a:tr>")
    return f'''<p:graphicFrame><p:nvGraphicFramePr><p:cNvPr id="{shape_id}" name="Table {shape_id}"/><p:cNvGraphicFramePr><a:graphicFrameLocks noGrp="1"/></p:cNvGraphicFramePr><p:nvPr/></p:nvGraphicFramePr><p:xfrm><a:off x="{x}" y="{y}"/><a:ext cx="{w}" cy="{h}"/></p:xfrm><a:graphic><a:graphicData uri="http://schemas.openxmlformats.org/drawingml/2006/table"><a:tbl><a:tblPr firstRow="1" bandRow="1"><a:tableStyleId>{{5940675A-B579-460E-94D1-54222C63F5DA}}</a:tableStyleId></a:tblPr><a:tblGrid>{grid}</a:tblGrid>{''.join(trs)}</a:tbl></a:graphicData></a:graphic></p:graphicFrame>'''


def editable_slide_xml(title, subtitle="", headers=None, rows=None):
    shapes = [
        ppt_text_box(2, 130000, 45000, SLIDE_W - 260000, 365000, title, 1900, True, None, "22A7D8"),
    ]
    if subtitle:
        shapes.append(ppt_text_box(3, 130000, 430000, SLIDE_W - 260000, 215000, subtitle, 900, False, None, BLACK, "r"))
    if headers and rows is not None:
        shapes.append(ppt_table(4, 130000, 670000, SLIDE_W - 260000, SLIDE_H - 800000, headers, rows))
    sp_tree = "".join(shapes)
    return f'''<?xml version="1.0" encoding="UTF-8" standalone="yes"?><p:sld xmlns:a="http://schemas.openxmlformats.org/drawingml/2006/main" xmlns:r="http://schemas.openxmlformats.org/officeDocument/2006/relationships" xmlns:p="http://schemas.openxmlformats.org/presentationml/2006/main"><p:cSld><p:spTree><p:nvGrpSpPr><p:cNvPr id="1" name=""/><p:cNvGrpSpPr/><p:nvPr/></p:nvGrpSpPr><p:grpSpPr><a:xfrm><a:off x="0" y="0"/><a:ext cx="0" cy="0"/><a:chOff x="0" y="0"/><a:chExt cx="0" cy="0"/></a:xfrm></p:grpSpPr>{sp_tree}</p:spTree></p:cSld><p:clrMapOvr><a:masterClrMapping/></p:clrMapOvr></p:sld>'''


def cover_slide_xml(title, subtitle):
    shapes = [
        ppt_text_box(2, 0, 0, SLIDE_W, 540000, "NORTHERN RAILWAY - MORADABAD DIVISION", 1350, True, NAVY, WHITE),
        ppt_text_box(3, 420000, 1180000, SLIDE_W - 840000, 760000, title, 2400, True, None, "22A7D8"),
        ppt_text_box(4, 860000, 2040000, SLIDE_W - 1720000, 430000, subtitle, 980, False, LIGHT, BLACK),
        ppt_text_box(5, 1350000, 3120000, 2400000, 720000, "Green\nWithin control", 920, True, GREEN_LIGHT, BLACK),
        ppt_text_box(6, 4860000, 3120000, 2400000, 720000, "Amber\nWatch range", 920, True, AMBER_LIGHT, BLACK),
        ppt_text_box(7, 8370000, 3120000, 2400000, 720000, "Red\nImmediate review", 920, True, RED_LIGHT, BLACK),
        ppt_text_box(8, 0, SLIDE_H - 420000, SLIDE_W, 420000, "Figures in thousands with Crore equivalent shown inside table cells", 900, False, BLUE, WHITE),
    ]
    sp_tree = "".join(shapes)
    return f'''<?xml version="1.0" encoding="UTF-8" standalone="yes"?><p:sld xmlns:a="http://schemas.openxmlformats.org/drawingml/2006/main" xmlns:r="http://schemas.openxmlformats.org/officeDocument/2006/relationships" xmlns:p="http://schemas.openxmlformats.org/presentationml/2006/main"><p:cSld><p:spTree><p:nvGrpSpPr><p:cNvPr id="1" name=""/><p:cNvGrpSpPr/><p:nvPr/></p:nvGrpSpPr><p:grpSpPr><a:xfrm><a:off x="0" y="0"/><a:ext cx="0" cy="0"/><a:chOff x="0" y="0"/><a:chExt cx="0" cy="0"/></a:xfrm></p:grpSpPr>{sp_tree}</p:spTree></p:cSld><p:clrMapOvr><a:masterClrMapping/></p:clrMapOvr></p:sld>'''


def split_section(title, headers, rows):
    slides = []
    col_count = len(headers)
    if col_count <= 6:
        max_rows = 18
    elif col_count <= 9:
        max_rows = 13
    else:
        max_rows = 10
    for r_idx in range(0, len(rows), max_rows):
        part_rows = rows[r_idx:r_idx + max_rows]
        slide_title = title if len(rows) <= max_rows else f"{title} (Rows {r_idx + 1}-{r_idx + len(part_rows)})"
        slides.append(editable_slide_xml(slide_title, "Figures in thousands.", headers, part_rows))
    return slides


def build_pptx_from_template(output_path, sections, subtitle):
    if not TEMPLATE_PPTX.exists():
        raise RuntimeError(f"Template PPTX not found: {TEMPLATE_PPTX}")
    slides = [cover_slide_xml("Moradabad Division Budget & FR Analysis", subtitle)]
    for title, headers, rows in sections:
        slides.extend(split_section(title, headers, rows))
    with zipfile.ZipFile(TEMPLATE_PPTX, "r") as src:
        content = src.read("[Content_Types].xml").decode("utf-8")
        pres = src.read("ppt/presentation.xml").decode("utf-8")
        pres_rels = src.read("ppt/_rels/presentation.xml.rels").decode("utf-8")
        slide_rel = src.read("ppt/slides/_rels/slide1.xml.rels")
        content = re.sub(r'<Override PartName="/ppt/slides/slide\d+\.xml" ContentType="application/vnd.openxmlformats-officedocument.presentationml.slide\+xml"/>', "", content)
        slide_overrides = "".join(f'<Override PartName="/ppt/slides/slide{i}.xml" ContentType="application/vnd.openxmlformats-officedocument.presentationml.slide+xml"/>' for i in range(1, len(slides) + 1))
        content = content.replace("</Types>", f"{slide_overrides}</Types>")
        sld_ids = "".join(f'<p:sldId id="{255 + i}" r:id="rId{i + 1}"/>' for i in range(1, len(slides) + 1))
        pres = re.sub(r"<p:sldIdLst>.*?</p:sldIdLst>", f"<p:sldIdLst>{sld_ids}</p:sldIdLst>", pres, flags=re.S)
        pres_rels = re.sub(r'<Relationship Id="rId\d+" Type="http://schemas.openxmlformats.org/officeDocument/2006/relationships/slide" Target="slides/slide\d+\.xml"/>', "", pres_rels)
        slide_rels = "".join(f'<Relationship Id="rId{i + 1}" Type="http://schemas.openxmlformats.org/officeDocument/2006/relationships/slide" Target="slides/slide{i}.xml"/>' for i in range(1, len(slides) + 1))
        pres_rels = pres_rels.replace("</Relationships>", f"{slide_rels}</Relationships>")
        with zipfile.ZipFile(output_path, "w", zipfile.ZIP_DEFLATED) as dst:
            for item in src.infolist():
                name = item.filename
                if name == "[Content_Types].xml" or name in {"ppt/presentation.xml", "ppt/_rels/presentation.xml.rels"}:
                    continue
                if re.match(r"ppt/slides/(?:_rels/)?slide\d+\.xml(?:\.rels)?$", name):
                    continue
                dst.writestr(item, src.read(name))
            dst.writestr("[Content_Types].xml", content)
            dst.writestr("ppt/presentation.xml", pres)
            dst.writestr("ppt/_rels/presentation.xml.rels", pres_rels)
            for i, xml in enumerate(slides, 1):
                dst.writestr(f"ppt/slides/slide{i}.xml", xml)
                dst.writestr(f"ppt/slides/_rels/slide{i}.xml.rels", slide_rel)


def build():
    payload = apply_completed_period(load_json_assignment(ROOT / "data" / "current_payload.js", "window.CURRENT_PAYLOAD"))
    fr = load_fr_data()
    fr_as_on = load_fr_as_on()
    current_basis = current_as_on_label()
    demand_cols = payload["demand"]["columns"]
    staff_cols = payload["staff"]["columns"]
    nonstaff_cols = payload["nonstaff"]["columns"]
    current_sections = [
        (f"Demand SMH Wise - {current_basis}", *table_from_payload(payload["demand"], demand_cols, payload["demand"]["rows"])),
        (f"PU Staff Current Year - {current_basis}", *table_from_payload(payload["staff"], staff_cols, payload["staff"]["rows"])),
        (f"PU Non Staff Current Year - {current_basis}", *table_from_payload(payload["nonstaff"], nonstaff_cols, payload["nonstaff"]["rows"])),
        (f"PU Previous Year Comparison - {current_basis}", *table_from_payload(payload["pu_prev"])),
        (f"Demand Previous Year Comparison - {current_basis}", *table_from_payload(payload["demand_prev"])),
    ]
    drm_staff_rows = filtered_pu_rows(payload["staff"]["rows"], ["01", "02", "03", "04", "07", "10", "11", "12", "13", "15", "16", "25"])
    drm_nonstaff_rows = filtered_pu_rows(payload["nonstaff"]["rows"], ["27", "28", "30", "32", "60"])
    drm_sections = [
        (f"Demand SMH Wise - {current_basis}", *table_from_payload(payload["demand"], demand_cols, payload["demand"]["rows"])),
        (f"PU Wise - Staff - {current_basis}", *table_from_payload(payload["staff"], staff_cols, drm_staff_rows)),
        (f"PU Wise - Non-Staff Part 1 - {current_basis}", *table_from_payload(payload["nonstaff"], nonstaff_cols, drm_nonstaff_rows)),
        (f"Open Line FR Report - As On {fr_as_on}", *fr_report_table(fr[0])),
        (f"Open Line FR Fund Wise - As On {fr_as_on}", *fr_fund_table(fr[0])),
    ]
    fr_sections = [
        (f"Open Line FR Report - As On {fr_as_on}", *fr_report_table(fr[0])),
        (f"Open Line FR Fund Wise - As On {fr_as_on}", *fr_fund_table(fr[0])),
        (f"GSU FR Report - As On {fr_as_on}", *fr_report_table(fr[1])),
        (f"GSU FR Fund Wise - As On {fr_as_on}", *fr_fund_table(fr[1])),
    ]
    OUT.mkdir(parents=True, exist_ok=True)
    write_excel(current_sections, CURRENT_XLSX)
    write_excel(fr_sections, FR_XLSX)
    write_excel(drm_sections, XLSX)
    write_pdf(current_sections, CURRENT_PDF)
    write_pdf(fr_sections, FR_PDF)
    build_pptx_from_template(CURRENT_PPTX, current_sections, "Accounts Dept | FY 2026-2027 | Current / Previous Year Budget Analysis | Completed JUL 2026")
    build_pptx_from_template(PPTX, drm_sections, "Accounts Dept | FY 2026-2027 | DRM Budget & FR Analysis | Completed JUL 2026")
    for path in (CURRENT_PPTX, PPTX):
        with zipfile.ZipFile(path) as z:
            assert z.testzip() is None
            assert "ppt/presentation.xml" in z.namelist()
    for path in (CURRENT_XLSX, FR_XLSX, XLSX):
        with zipfile.ZipFile(path) as z:
            assert z.testzip() is None
    for path in (CURRENT_PDF, FR_PDF):
        assert path.read_bytes().startswith(b"%PDF")
    print(f"Generated {CURRENT_XLSX}")
    print(f"Generated {CURRENT_PDF}")
    print(f"Generated {FR_XLSX}")
    print(f"Generated {FR_PDF}")
    print(f"Generated {XLSX}")
    print(f"Generated {CURRENT_PPTX}")
    print(f"Generated {PPTX}")


if __name__ == "__main__":
    build()
