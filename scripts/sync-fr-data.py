from datetime import datetime
from pathlib import Path
import filecmp
import json
import os
import re
import shutil
import stat
import subprocess
import sys

from openpyxl import load_workbook

REPO_ROOT = Path(__file__).resolve().parents[1]
FR_ROOT = REPO_ROOT / "data" / "fr"
FR_TARGET = FR_ROOT / "FR_Budget_Status.xlsx"
FR_MANIFEST = FR_ROOT / "fr-upload-manifest.json"
FR_PAGE = REPO_ROOT / "pages" / "fr.html"
FUNDS = ["CAP", "DRF", "DF", "CAPITAL FUND", "S FUND", "EBR-(S)", "RRSK"]
FUND_STARTS = {
    "CAP": 2,
    "DRF": 5,
    "DF": 8,
    "CAPITAL FUND": 11,
    "S FUND": 14,
    "EBR-(S)": 17,
    "RRSK": 20,
    "TOTAL": 23,
}


def rel(path):
    return str(path.relative_to(REPO_ROOT)).replace(os.sep, "/")


def number(value):
    if isinstance(value, (int, float)):
        return float(value)
    if value is None:
        return 0.0
    text = str(value).replace(",", "").strip()
    if not text or text.startswith("#"):
        return 0.0
    try:
        return float(text)
    except ValueError:
        return 0.0


def clean(value):
    return str(value or "").strip()


def fund_values(row, start):
    sba = number(row[start] if start < len(row) else 0)
    ae = number(row[start + 1] if start + 1 < len(row) else 0)
    workbook_var = number(row[start + 2] if start + 2 < len(row) else ae - sba)
    available = sba - ae
    spent = (ae / sba * 100) if sba else 0
    remain = (available / sba * 100) if sba else 0
    return {
        "sba": sba,
        "ae": ae,
        "var": workbook_var,
        "available": available,
        "spentPct": spent,
        "remainPct": remain,
        "workbookVar": workbook_var,
    }


def parse_sheet(ws):
    title = clean(ws.cell(1, 1).value)
    records = []
    total_row = None
    for row in ws.iter_rows(min_row=4, max_col=27, values_only=True):
        plan_head = clean(row[0])
        plan_name = clean(row[1])
        if not plan_head and not plan_name:
            continue
        if plan_head.lower() == "total":
            total_row = row
            break
        if plan_name.lower().startswith("copy forwarded"):
            break
        funds = {fund: fund_values(row, FUND_STARTS[fund]) for fund in FUNDS}
        funds["TOTAL"] = fund_values(row, FUND_STARTS["TOTAL"])
        records.append({"planHead": plan_head, "planName": plan_name, "funds": funds})
    if total_row is None:
        total_funds = {}
        for fund in [*FUNDS, "TOTAL"]:
            total_funds[fund] = {
                "sba": sum(rec["funds"][fund]["sba"] for rec in records),
                "ae": sum(rec["funds"][fund]["ae"] for rec in records),
                "var": sum(rec["funds"][fund]["workbookVar"] for rec in records),
                "workbookVar": sum(rec["funds"][fund]["workbookVar"] for rec in records),
                "available": 0,
                "spentPct": 0,
                "remainPct": 0,
            }
            total_funds[fund]["available"] = total_funds[fund]["sba"] - total_funds[fund]["ae"]
            total_funds[fund]["spentPct"] = total_funds[fund]["ae"] / total_funds[fund]["sba"] * 100 if total_funds[fund]["sba"] else 0
            total_funds[fund]["remainPct"] = total_funds[fund]["available"] / total_funds[fund]["sba"] * 100 if total_funds[fund]["sba"] else 0
    else:
        total_funds = {fund: fund_values(total_row, FUND_STARTS[fund]) for fund in FUNDS}
        total_funds["TOTAL"] = fund_values(total_row, FUND_STARTS["TOTAL"])
    return {
        "sheetName": ws.title,
        "title": title,
        "records": records,
        "total": {"planHead": "Total", "planName": "", "funds": total_funds},
        "fundAnalysis": {fund: total_funds[fund] for fund in FUNDS},
    }


def parse_workbook(path):
    workbook = load_workbook(path, data_only=True, read_only=True)
    return [parse_sheet(sheet) for sheet in workbook.worksheets[:2]]


def as_on_from_data(workbook_data, fallback_path):
    text = " ".join(sheet.get("title", "") for sheet in workbook_data)
    match = re.search(r"as\s+on\s+([0-9]{1,2}[.][0-9]{1,2}[.][0-9]{4})", text, re.I)
    if match:
        return match.group(1)
    match = re.search(r"([0-9]{1,2})[.]([0-9]{1,2})[.]([0-9]{4})", fallback_path.name)
    return ".".join(match.groups()) if match else ""


def iso_date_from_ddmmyyyy(value):
    try:
        return datetime.strptime(value, "%d.%m.%Y").date().isoformat()
    except ValueError:
        return value


def backup_listing():
    root = FR_ROOT / "backups"
    if not root.exists():
        return []
    backups = []
    for folder in sorted([p for p in root.iterdir() if p.is_dir()], key=lambda p: p.name, reverse=True)[:2]:
        backups.append({"name": folder.name, "files": [rel(p) for p in sorted(folder.glob("FR_Budget_Status.*")) if p.is_file()]})
    return backups


def keep_two_backups():
    root = FR_ROOT / "backups"
    if not root.exists():
        return
    backups = sorted([p for p in root.iterdir() if p.is_dir()], key=lambda p: p.name, reverse=True)
    for old in backups[2:]:
        shutil.rmtree(old, onerror=remove_readonly)


def remove_readonly(func, path, _exc):
    os.chmod(path, stat.S_IWRITE)
    func(path)


def update_fr_page(workbook_data, source_name, data_as_on, uploaded_at):
    text = FR_PAGE.read_text(encoding="utf-8")
    data_json = json.dumps(workbook_data, separators=(",", ":"))
    text = re.sub(r"const\s+workbookData\s*=\s*\[.*?\];\s*const\s+funds", f"const workbookData = {data_json};\n    const funds", text, flags=re.S)
    text = re.sub(r"Data as on <strong>.*?</strong>", f"Data as on <strong>{data_as_on}</strong>", text)
    text = re.sub(r"Source <strong>.*?</strong>", f"Source <strong>{source_name}</strong>", text)
    text = re.sub(r"Stored <strong>.*?</strong>", f"Stored <strong>{uploaded_at}</strong>", text)
    FR_PAGE.write_text(text, encoding="utf-8")


def refresh_exports():
    script = REPO_ROOT / "scripts" / "generate_drm_exports.py"
    if not script.exists():
        print("Export refresh skipped: generator not found")
        return
    result = subprocess.run([sys.executable, str(script)], cwd=REPO_ROOT, text=True, capture_output=True)
    if result.stdout.strip():
        print(result.stdout.strip())
    if result.returncode:
        if result.stderr.strip():
            print(result.stderr.strip())
        raise RuntimeError("Export refresh failed")


def main():
    if len(sys.argv) < 2:
        raise SystemExit("Usage: python scripts\\sync-fr-data.py <FR workbook path> [display source name]")
    source = Path(sys.argv[1]).resolve()
    source_display_name = sys.argv[2] if len(sys.argv) > 2 else source.name
    if not source.exists():
        raise SystemExit(f"FR workbook not found: {source}")
    workbook_data = parse_workbook(source)
    data_as_on = as_on_from_data(workbook_data, source)
    uploaded_at = datetime.now().isoformat(timespec="seconds")
    display_uploaded_at = datetime.now().strftime("%d.%m.%Y %H:%M:%S")

    FR_ROOT.mkdir(parents=True, exist_ok=True)
    backup_name = ""
    if FR_TARGET.exists() and not filecmp.cmp(source, FR_TARGET, shallow=False):
        backup_name = datetime.now().strftime("%Y%m%d-%H%M%S")
        backup_dir = FR_ROOT / "backups" / backup_name
        backup_dir.mkdir(parents=True, exist_ok=True)
        shutil.copy2(FR_TARGET, backup_dir / FR_TARGET.name)
    if not FR_TARGET.exists() or not filecmp.cmp(source, FR_TARGET, shallow=False):
        shutil.copy2(source, FR_TARGET)
    keep_two_backups()
    manifest = {
        "uploadedAt": uploaded_at,
        "dataAsOn": iso_date_from_ddmmyyyy(data_as_on),
        "activeFile": rel(FR_TARGET),
        "originalName": source_display_name,
        "backup": backup_name,
        "backups": backup_listing(),
    }
    FR_MANIFEST.write_text(json.dumps(manifest, indent=2), encoding="utf-8")
    update_fr_page(workbook_data, source_display_name, data_as_on, display_uploaded_at)
    refresh_exports()
    print(f"FR sync complete: {source_display_name}")
    print(f"Data as on: {data_as_on}")
    print(f"Backup: {backup_name or 'not needed'}")
    for sheet in workbook_data:
        total = sheet["total"]["funds"]["TOTAL"]
        print(f"{sheet['sheetName']}: SBA={total['sba']:.0f} AE={total['ae']:.0f} Available={total['available']:.0f} Exp%={total['spentPct']:.2f}")


if __name__ == "__main__":
    main()
