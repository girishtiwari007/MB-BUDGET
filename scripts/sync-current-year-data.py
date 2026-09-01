from copy import deepcopy
import argparse
from datetime import datetime
from pathlib import Path
import importlib.util
import json
import re
import shutil
import sys

import openpyxl
import xlrd

from export_refresh import refresh_exports


REPO_ROOT = Path(__file__).resolve().parents[1]
YEAR = "2026-2027"
FY = "2026-27"
PREVIOUS_YEAR = "2025-2026"
PREVIOUS_FY = "2025-26"
DEFAULT_SOURCE = Path(r"C:\Users\HP\Downloads\PORTAL DATA")
MONTHS = ["APR", "MAY", "JUN", "JUL", "AUG", "SEP", "OCT", "NOV", "DEC", "JAN", "FEB", "MAR"]
STAFF_CODES = {"01", "02", "03", "04", "07", "08", "10", "11", "12", "13", "14", "15", "16", "17", "20", "25", "26", "29", "34", "39", "40", "42", "43", "44", "53", "54", "63"}
SOURCE_FILES = {
    "currPuBudget": ("PU-BUDGET.xls", "pu-budget.xls"),
    "currPuMonth": ("PU-MONTH-ACTUAL.xls", "pu-month-actual.xls"),
    "currPuDeptDemandSmhBudget": ("PU-DEPT-DEMAND-SMH-BUDGET.xls", "pu-dept-demand-smh-budget.xls"),
    "currPuDeptDemandSmhActual": ("PU-DEPT-DEMAND-SMH-ACTUAL.xls", "pu-dept-demand-smh-actual.xls"),
    "currSmhBudget": ("DEMAND-SMH-BUGDET.xls", "demand-smh-budget.xls"),
    "currSmhMonth": ("DEMAND-SMH-ACTUAL.xls", "demand-smh-actual.xls"),
}
DEMAND_DEPARTMENT = {
    "03": "PERSONNEL / STORE And Office Staff",
    "04": "ENGINEERING / PWAY",
    "05": "Mechanical LOCO Shed Roza",
    "06": "Electrical General / Mech C&W",
    "07": "S&T / TRD",
    "08": "MECHANICAL / Running Staff",
    "09": "OPERATING / Commercial",
    "10": "Operating Expenses - Fuel / Traction",
    "11": "MEDICAL",
    "12": "SECURITY",
    "13": "Pension and Retirement",
    "12N": "Suspense Heads",
}


def load_upload_helpers():
    module_path = REPO_ROOT / "scripts" / "local-upload-server.py"
    spec = importlib.util.spec_from_file_location("local_upload_server", module_path)
    module = importlib.util.module_from_spec(spec)
    spec.loader.exec_module(module)
    return module


def read_rows(path):
    if path.suffix.lower() == ".xlsx":
        workbook = openpyxl.load_workbook(path, read_only=True, data_only=True)
        sheet = workbook[workbook.sheetnames[0]]
        return [[cell for cell in row] for row in sheet.iter_rows(values_only=True)]
    workbook = xlrd.open_workbook(str(path))
    sheet = workbook.sheet_by_index(0)
    return [sheet.row_values(index) for index in range(sheet.nrows)]


def clean(value):
    return re.sub(r"\s+", " ", str(value or "")).strip().upper()


def number(value):
    try:
        if value in ("", None):
            return 0
        return float(value)
    except Exception:
        return 0


def workbook_table(path):
    raw = [row for row in read_rows(path) if any(str(cell).strip() for cell in row)]
    header_index = next((idx for idx, row in enumerate(raw) if clean(row[0]) == "AU"), -1)
    if header_index < 0:
        raise RuntimeError(f"AU header row not found in {path.name}")
    return {
        "path": path,
        "headers": [clean(cell) for cell in raw[header_index]],
        "rows": raw[header_index + 1:],
    }


def col_index(headers, needles):
    wanted = [clean(item) for item in needles]
    for idx, header in enumerate(headers):
        if all(item in header for item in wanted):
            return idx
    raise RuntimeError("Missing column: " + " ".join(needles))


def month_count(month):
    key = clean(month)[:3]
    return MONTHS.index(key) + 1 if key in MONTHS else 4


def period_key(value):
    match = re.search(r"([A-Z]{3})\s*(20\d{2})", clean(value))
    return f"{match.group(1)} {match.group(2)}" if match else ""


def period_from_label(label, idx=-1):
    key = period_key(label)
    if not key:
        return None
    month, year = key.split()
    return {"idx": idx, "month": month, "year": int(year), "count": month_count(month), "label": key}


def override_period(matches, label, role, allow_missing=False):
    if not label:
        return None
    key = period_key(label)
    found = next((item for item in matches if item["label"].upper() == key), None)
    if not found:
        if allow_missing:
            return period_from_label(label)
        available = ", ".join(item["label"] for item in matches) or "none"
        raise RuntimeError(f"{role} month {label} is not available in uploaded actual columns. Available: {available}")
    return found


def actual_periods(headers, completed_label=None, running_label=None):
    matches = []
    for idx, header in enumerate(headers):
        match = re.search(r"ACTUALS\s+UPTO\s+([A-Z]{3})\s+(20\d{2})", header)
        if match:
            matches.append({"idx": idx, "month": match.group(1), "year": int(match.group(2)), "count": month_count(match.group(1)), "label": f"{match.group(1)} {match.group(2)}"})
    matches.sort(key=lambda item: (item["year"], item["count"]))
    completed_override = override_period(matches, completed_label, "Completed actual")
    running_override = override_period(matches, running_label, "Running", allow_missing=True)
    if completed_override:
        completed = completed_override
        running = running_override or next((item for item in matches if (item["year"], item["count"]) > (completed["year"], completed["count"])), completed)
        if (running["year"], running["count"]) < (completed["year"], completed["count"]):
            raise RuntimeError(f"Running month {running['label']} cannot be before completed month {completed['label']}.")
        return completed, running
    if running_override and len(matches) >= 2:
        completed = matches[matches.index(running_override) - 1] if matches.index(running_override) > 0 else running_override
        return completed, running_override
    if len(matches) >= 2:
        return matches[-2], matches[-1]
    if matches:
        return matches[-1], matches[-1]
    return {"idx": -1, "month": "AUG", "year": 2026, "count": 5, "label": "AUG 2026"}, {"idx": -1, "month": "SEP", "year": 2026, "count": 6, "label": "SEP 2026"}


def available_actual_periods(headers):
    periods = []
    seen = set()
    for idx, header in enumerate(headers):
        match = re.search(r"ACTUALS\s+UPTO\s+([A-Z]{3})\s+(20\d{2})", header)
        if match:
            label = f"{match.group(1)} {match.group(2)}"
            if label not in seen:
                seen.add(label)
                periods.append({"idx": idx, "month": match.group(1), "year": int(match.group(2)), "count": month_count(match.group(1)), "label": label})
    periods.sort(key=lambda item: (item["year"], item["count"]))
    return periods


def source_file(source_root, role):
    source_root = Path(source_root).resolve()
    source_name, target_name = SOURCE_FILES[role]
    source_path = source_root / source_name
    if not source_path.exists():
        source_path = source_root / target_name
    return source_path


def available_period_labels(source_root):
    path = source_file(source_root, "currSmhBudget")
    if not path.exists():
        raise RuntimeError(f"Current-year budget file not found for month sensing: {path.name}")
    table = workbook_table(path)
    return [item["label"] for item in available_actual_periods(table["headers"])]


def find_bp(headers, period):
    exact = [idx for idx, header in enumerate(headers) if "BP" in header and "UPTO" in header and period["month"] in header and str(period["year"]) in header]
    return exact[-1] if exact else -1


def find_coppy(headers, period):
    previous_year = str(period["year"] - 1)
    exact = [idx for idx, header in enumerate(headers) if "COPPY" in header and "UPTO" in header and period["month"] in header and previous_year in header]
    return exact[-1] if exact else -1


def code_from_label(label, prefix):
    match = re.search(prefix + r"\s*-\s*([0-9A-Z]+)", str(label or ""), re.I)
    return match.group(1).upper() if match else ""


def demand_from_smh(label):
    code = code_from_label(label, "SMH")
    match = re.match(r"^(\d+)([A-Z]*)$", code)
    if match:
        return f"Demand {int(match.group(1)) + 2:02d}{match.group(2)} / {code}"
    return code or str(label or "")


def demand_key(label):
    match = re.search(r"Demand\s+([0-9A-Z]+)", str(label or ""), re.I)
    return match.group(1).upper() if match else ""


def with_department(row):
    row["Department"] = DEMAND_DEPARTMENT.get(demand_key(row["Name"]), "")
    return row


def is_suspense(row):
    return bool(re.search(r"\b(12N|10N)\b", str(row.get("Name", "")), re.I)) or "SUSPENSE" in str(row.get("Department", "")).upper()


def summary_row(label, oba, ae, months, bp):
    bp = number(bp) if bp not in (None, "") else (number(oba) / 12 * months)
    oba = number(oba)
    ae = number(ae)
    return {
        "Name": label,
        "OBA": round(oba),
        "BP": round(bp),
        "AE": round(ae),
        "Variation": round(ae - bp),
        "BPPercent": (ae / bp * 100) if bp else 0,
        "Remaining": round(oba - ae),
        "OBAPercent": (ae / oba * 100) if oba else 0,
        "Months": months,
    }


def previous_row(label, previous_oba, current_oba, ae_current, ae_previous, months):
    previous_oba = number(previous_oba)
    current_oba = number(current_oba)
    ae_current = number(ae_current)
    ae_previous = number(ae_previous)
    previous_bp = previous_oba / 12 * months
    bp = current_oba / 12 * months
    return {
        "Name": label,
        "PreviousOBA": round(previous_oba),
        "PreviousBP": previous_bp,
        "AEPrevious": round(ae_previous),
        "OBA": round(current_oba),
        "BP": bp,
        "AECurrent": round(ae_current),
        "VariationBP": ae_current - bp,
        "BPPercent": (ae_current / bp * 100) if bp else 0,
        "VariationActual": round(ae_current - ae_previous),
        "OBAPercent": (ae_current / current_oba * 100) if current_oba else 0,
        "Months": months,
    }


def add_total(rows, previous=False):
    normal = [row for row in rows if not is_suspense(row)]
    suspense = [row for row in rows if is_suspense(row)]
    months = normal[0].get("Months", 4) if normal else 4
    if previous:
        total = previous_row(
            "Total",
            sum(number(row.get("PreviousOBA")) for row in normal),
            sum(number(row.get("OBA")) for row in normal),
            sum(number(row.get("AECurrent")) for row in normal),
            sum(number(row.get("AEPrevious")) for row in normal),
            months,
        )
    else:
        total = summary_row(
            "Total",
            sum(number(row.get("OBA")) for row in normal),
            sum(number(row.get("AE")) for row in normal),
            months,
            sum(number(row.get("BP")) for row in normal),
        )
    return normal + [total] + suspense


def build_current(table, field, first_label, title, demand=False, completed_month=None, running_month=None):
    name_idx = col_index(table["headers"], [field])
    oba_idx = col_index(table["headers"], ["BG_ISL", "2026-2027"])
    completed, running = actual_periods(table["headers"], completed_month, running_month)
    ae_idx = completed["idx"]
    rows = []
    for raw in table["rows"]:
        name = str(raw[name_idx] if name_idx < len(raw) else "").strip()
        if not name or clean(name) == "TOTAL":
            continue
        label = demand_from_smh(name) if demand else name
        row = summary_row(label, raw[oba_idx], raw[ae_idx] if ae_idx >= 0 else 0, completed["count"], None)
        if demand:
            with_department(row)
        rows.append(row)
    columns = [
        {"key": "Name", "label": first_label, "format": "text"},
        *([{"key": "Department", "label": "Department", "format": "text"}] if demand else []),
        {"key": "OBA", "label": "A\nOBA\nBG_ISL 2026-27", "format": "money"},
        {"key": "BP", "label": f"B\nBP\nA / 12 * {completed['count']}", "format": "money"},
        {"key": "AE", "label": f"C\nAE\nActuals up to {completed['label']}", "format": "money"},
        {"key": "Variation", "label": "D\nVariation\nC - B", "format": "money"},
        {"key": "BPPercent", "label": "E\n% BP\nC / B", "format": "int"},
        {"key": "Remaining", "label": "F\nBudget Remaining\nA - C", "format": "money"},
        {"key": "OBAPercent", "label": "G\n% OBA Utilized\nC / A", "format": "int"},
    ]
    return {"title": title, "columns": columns, "rows": add_total(rows)}, completed, running


def map_column(table, field, needles, labeler=lambda value: value):
    name_idx = col_index(table["headers"], [field])
    value_idx = col_index(table["headers"], needles)
    result = {}
    for raw in table["rows"]:
        name = str(raw[name_idx] if name_idx < len(raw) else "").strip()
        if name and clean(name) != "TOTAL":
            result[labeler(name)] = number(raw[value_idx] if value_idx < len(raw) else 0)
    return result


def build_previous(prev_budget, curr_budget, field, first_label, title, demand=False, completed_month=None, running_month=None):
    labeler = demand_from_smh if demand else (lambda value: value)
    rg = map_column(prev_budget, field, ["RG", "2025-2026"], labeler)
    bg = map_column(curr_budget, field, ["BG_ISL", "2026-2027"], labeler)
    name_idx = col_index(curr_budget["headers"], [field])
    completed, _running = actual_periods(curr_budget["headers"], completed_month, running_month)
    ae_idx = completed["idx"]
    coppy_idx = find_coppy(curr_budget["headers"], completed)
    current = {}
    previous = {}
    for raw in curr_budget["rows"]:
        name = str(raw[name_idx] if name_idx < len(raw) else "").strip()
        if name and clean(name) != "TOTAL":
            label = labeler(name)
            current[label] = number(raw[ae_idx] if ae_idx >= 0 and ae_idx < len(raw) else 0)
            previous[label] = number(raw[coppy_idx] if coppy_idx >= 0 and coppy_idx < len(raw) else 0)
    previous_label = f"{completed['month']} {completed['year'] - 1}"
    rows = []
    for label in rg:
        row = previous_row(label, rg.get(label, 0), bg.get(label, 0), current.get(label, 0), previous.get(label, 0), completed["count"])
        if demand:
            with_department(row)
        rows.append(row)
    columns = [
        {"key": "Name", "label": first_label, "format": "text"},
        *([{"key": "Department", "label": "Department", "format": "text"}] if demand else []),
        {"key": "PreviousOBA", "label": "A\nPrevious OBA\nRG 2025-26", "format": "money"},
        {"key": "PreviousBP", "label": f"B\nPrevious Budget Proportion\nA / 12 * {completed['count']}", "format": "money"},
        {"key": "AEPrevious", "label": f"C\nPrevious Actual Expenditure\nup to {previous_label}", "format": "money"},
        {"key": "OBA", "label": "D\nCurrent OBA\nBG_ISL 2026-27", "format": "money"},
        {"key": "BP", "label": f"E\nCurrent Budget Proportion\nD / 12 * {completed['count']}", "format": "money"},
        {"key": "AECurrent", "label": f"F\nCurrent Actual Expenditure\nup to {completed['label']}", "format": "money"},
        {"key": "VariationBP", "label": "G\nBudget Variation\nF - E", "format": "money"},
        {"key": "BPPercent", "label": "H\nCurrent Budget Proportion %\nF / E", "format": "int"},
        {"key": "VariationActual", "label": "I\nActual Expenditure Variation\nF - C", "format": "money"},
        {"key": "OBAPercent", "label": "J\nCurrent OBA Utilization %\nF / D", "format": "int"},
    ]
    return {"title": title, "columns": columns, "rows": add_total(rows, True)}


def monthly_map(table, field, labeler=lambda value: value):
    name_idx = col_index(table["headers"], [field])
    indices = []
    for month in MONTHS:
        indices.append(next((idx for idx, header in enumerate(table["headers"]) if header.startswith(month + " ")), -1))
    out = {}
    for raw in table["rows"]:
        name = str(raw[name_idx] if name_idx < len(raw) else "").strip()
        if name and clean(name) != "TOTAL":
            out[labeler(name)] = [number(raw[idx] if idx >= 0 and idx < len(raw) else 0) for idx in indices]
    return out


def update_reports_data(pu_budget, smh_budget, pu_month, smh_month, dept_month, current_payload, status_as_on):
    json_path = REPO_ROOT / "data" / "reports-data.json"
    reports = json.loads(json_path.read_text(encoding="utf-8-sig"))
    for row in current_payload["demand"]["rows"]:
        if row["Name"] == "Total":
            continue
        reports.setdefault("budget", {}).setdefault("demand", {}).setdefault(row["Name"], {})[FY] = {"oba": row["OBA"], "ae": row["AE"], "bp": row["BP"]}
    for key in ("staff", "nonstaff"):
        for row in current_payload[key]["rows"]:
            if row["Name"] == "Total":
                continue
            reports.setdefault("budget", {}).setdefault("pu", {}).setdefault(row["Name"], {})[FY] = {"oba": row["OBA"], "ae": row["AE"], "bp": row["BP"]}
    for label, values in monthly_map(smh_month, "SMH", demand_from_smh).items():
        reports.setdefault("monthly", {}).setdefault("demand", {}).setdefault(label, {})[FY] = values
    for label, values in monthly_map(pu_month, "PUCODE").items():
        reports.setdefault("monthly", {}).setdefault("pu", {}).setdefault(label, {})[FY] = values
    dept_idx = col_index(dept_month["headers"], ["DEPARTMENTCODE"])
    dept_month_indices = [next((idx for idx, header in enumerate(dept_month["headers"]) if header.startswith(month + " ")), -1) for month in MONTHS]
    dept_totals = {}
    for raw in dept_month["rows"]:
        dept = str(raw[dept_idx] if dept_idx < len(raw) else "").strip() or "Not Classified"
        dept_totals.setdefault(dept, [0] * len(MONTHS))
        for pos, idx in enumerate(dept_month_indices):
            dept_totals[dept][pos] += number(raw[idx] if idx >= 0 and idx < len(raw) else 0)
    for dept, values in dept_totals.items():
        reports.setdefault("monthly", {}).setdefault("dept", {}).setdefault(dept, {})[FY] = values
    reports["generatedAt"] = status_as_on
    reports["statusAsOn"] = status_as_on
    json_path.write_text(json.dumps(reports, indent=2), encoding="utf-8")
    (REPO_ROOT / "data" / "reports-data.js").write_text("window.REPORTS_DATA = " + json.dumps(reports, indent=2) + ";\n", encoding="utf-8")


def write_current_payload(payload, completed, running, source_root, backup_name, basis_source):
    now = datetime.now().astimezone().isoformat(timespec="seconds")
    meta = {
        "statusAsOn": now,
        "sourceFolder": str(source_root),
        "financialYear": YEAR,
        "runningMonth": running["label"],
        "completedMonth": completed["label"],
        "basisSource": basis_source,
        "updatedAt": now,
        "backup": backup_name,
    }
    text = "window.CURRENT_PAYLOAD_META = " + json.dumps(meta) + ";\nwindow.CURRENT_PAYLOAD = " + json.dumps(payload, separators=(",", ":")) + ";\n"
    (REPO_ROOT / "data" / "current_payload.js").write_text(text, encoding="utf-8")
    return meta


def validate_current_payload(payload, completed, running):
    errors = []
    months = completed["count"]
    completed_label = completed["label"]
    running_label = running["label"]
    for key in ("demand", "staff", "nonstaff"):
        section = payload.get(key) or {}
        labels = " ".join(str(col.get("label", "")) for col in section.get("columns", []))
        if f"A / 12 * {months}" not in labels:
            errors.append(f"{key} BP column does not show A / 12 * {months}.")
        if completed_label not in labels:
            errors.append(f"{key} columns do not show completed month {completed_label}.")
        if running_label in labels:
            errors.append(f"{key} default calculation columns still include running month {running_label}.")
        for row in section.get("rows", []):
            if row.get("Name") == "Total" or is_suspense(row):
                continue
            oba = number(row.get("OBA"))
            expected_bp = round(oba / 12 * months)
            actual_bp = round(number(row.get("BP")))
            if abs(actual_bp - expected_bp) > 1:
                errors.append(f"{key} BP mismatch for {row.get('Name')}: {actual_bp} != {expected_bp}")
                break
            if row.get("Months") != months:
                errors.append(f"{key} month count mismatch for {row.get('Name')}: {row.get('Months')} != {months}")
                break
    for key in ("pu_prev", "demand_prev"):
        section = payload.get(key) or {}
        labels = " ".join(str(col.get("label", "")) for col in section.get("columns", []))
        if f"A / 12 * {months}" not in labels or f"D / 12 * {months}" not in labels:
            errors.append(f"{key} comparison BP columns do not use completed month count {months}.")
        for row in section.get("rows", []):
            if row.get("Name") == "Total" or is_suspense(row):
                continue
            previous_bp = round(number(row.get("PreviousOBA")) / 12 * months)
            current_bp = round(number(row.get("OBA")) / 12 * months)
            if abs(round(number(row.get("PreviousBP"))) - previous_bp) > 1:
                errors.append(f"{key} previous BP mismatch for {row.get('Name')}.")
                break
            if abs(round(number(row.get("BP"))) - current_bp) > 1:
                errors.append(f"{key} current BP mismatch for {row.get('Name')}.")
                break
    if errors:
        raise RuntimeError("Current-year calculation validation failed:\n- " + "\n- ".join(errors))


def copy_sources(source_root):
    source_root = Path(source_root).resolve()
    if not source_root.exists():
        raise RuntimeError(f"Source folder not found: {source_root}")
    missing = [src for src, target in SOURCE_FILES.values() if not (source_root / src).exists() and not (source_root / target).exists()]
    if missing:
        raise RuntimeError("Missing source files: " + ", ".join(missing))
    helpers = load_upload_helpers()
    year_dir = REPO_ROOT / "data" / "source-files" / YEAR
    year_dir.mkdir(parents=True, exist_ok=True)
    existing = [year_dir / target for _src, target in SOURCE_FILES.values() if (year_dir / target).exists()]
    backup_name = ""
    if existing:
        backup_name = datetime.now().strftime("%Y%m%d-%H%M%S")
        backup_dir = year_dir / "backups" / backup_name
        backup_dir.mkdir(parents=True, exist_ok=True)
        for file_path in existing:
            shutil.copy2(file_path, backup_dir / file_path.name)
    for _role, (source_name, target_name) in SOURCE_FILES.items():
        source_path = source_root / source_name
        if not source_path.exists():
            source_path = source_root / target_name
        if source_path.resolve() != (year_dir / target_name).resolve():
            shutil.copy2(source_path, year_dir / target_name)
    helpers.keep_two_backups(year_dir / "backups")
    return year_dir, backup_name, helpers


def sync_current_year(source_root=DEFAULT_SOURCE, refresh=True, completed_month=None, running_month=None):
    year_dir, backup_name, helpers = copy_sources(source_root)
    pu_budget = workbook_table(year_dir / "pu-budget.xls")
    smh_budget = workbook_table(year_dir / "demand-smh-budget.xls")
    pu_month = workbook_table(year_dir / "pu-month-actual.xls")
    smh_month = workbook_table(year_dir / "demand-smh-actual.xls")
    dept_month = workbook_table(year_dir / "pu-dept-demand-smh-actual.xls")
    prev_pu_budget = workbook_table(REPO_ROOT / "data" / "source-files" / PREVIOUS_YEAR / "pu-budget.xls")
    prev_smh_budget = workbook_table(REPO_ROOT / "data" / "source-files" / PREVIOUS_YEAR / "demand-smh-budget.xls")
    demand, completed, running = build_current(smh_budget, "SMH", "Demand No. / SMH-Grant", "Demand / SMH Wise Current Year", True, completed_month, running_month)
    pu_current, _completed, _running = build_current(pu_budget, "PUCODE", "PU", "PU Wise Current Year", False, completed_month, running_month)
    detail_rows = [row for row in pu_current["rows"] if row["Name"] != "Total"]
    staff = {"title": "PU Staff Current Year", "columns": deepcopy(pu_current["columns"]), "rows": add_total([row for row in detail_rows if code_from_label(row["Name"], "PU") in STAFF_CODES])}
    nonstaff = {"title": "PU Non-Staff Current Year", "columns": deepcopy(pu_current["columns"]), "rows": add_total([row for row in detail_rows if code_from_label(row["Name"], "PU") not in STAFF_CODES])}
    payload = {
        "demand": demand,
        "staff": staff,
        "nonstaff": nonstaff,
        "pu_prev": build_previous(prev_pu_budget, pu_budget, "PUCODE", "PU", "PU Wise Previous Year Comparison", False, completed_month, running_month),
        "demand_prev": build_previous(prev_smh_budget, smh_budget, "SMH", "Demand No. / SMH-Grant", "Demand / SMH Wise Previous Year Comparison", True, completed_month, running_month),
    }
    validate_current_payload(payload, completed, running)
    basis_source = "manual override" if completed_month or running_month else "auto-sensed from uploaded file"
    meta = write_current_payload(payload, completed, running, Path(source_root).resolve(), backup_name, basis_source)
    manifest = helpers.write_current_manifest(YEAR, str(Path(source_root).resolve()), backup_name)
    manifest.update({
        "uploadedAt": meta["updatedAt"],
        "statusAsOn": meta["statusAsOn"],
        "runningMonth": running["label"],
        "completedMonth": completed["label"],
        "basisSource": basis_source,
        "sourceFolder": str(Path(source_root).resolve()),
        "backup": backup_name,
    })
    manifest_path = REPO_ROOT / "data" / "source-files" / YEAR / "upload-manifest.json"
    manifest_path.write_text(json.dumps(manifest, indent=2), encoding="utf-8")
    update_reports_data(pu_budget, smh_budget, pu_month, smh_month, dept_month, payload, meta["statusAsOn"])
    export_log = refresh_exports("current-year-gui-sync") if refresh else ""
    return {
        "ok": True,
        "sourceFolder": str(Path(source_root).resolve()),
        "backup": backup_name,
        "completedMonth": completed["label"],
        "runningMonth": running["label"],
        "statusAsOn": meta["statusAsOn"],
        "manifest": manifest,
        "exportLog": export_log,
    }


def main():
    parser = argparse.ArgumentParser(description="Sync MB-BUDGET current-year data into the repository.")
    parser.add_argument("source", nargs="?", default=str(DEFAULT_SOURCE), help="Folder containing the six current-year source files.")
    parser.add_argument("--completed-month", default="", help="Override completed actual month, for example AUG 2026.")
    parser.add_argument("--running-month", default="", help="Override running month, for example SEP 2026.")
    args = parser.parse_args()
    source = Path(args.source).resolve()
    result = sync_current_year(source, completed_month=args.completed_month or None, running_month=args.running_month or None)
    print(json.dumps(result, indent=2))


if __name__ == "__main__":
    main()
